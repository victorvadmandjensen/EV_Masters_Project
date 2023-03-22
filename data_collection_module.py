from openpyxl import Workbook

class Data_Collection_Module:
    def __init__(self):
        self.counter = 1
        self.workbook = None
        self.ws = None
    
    def initialize_workbook(self):
        self.workbook = Workbook()
        self.ws = self.workbook.active
        self.ws.title = "Energy sent to battery"
        self.ws.cell(column=1,row=1).value = "Round"
        self.ws.cell(column=2,row=1).value = "Energy sent to battery by players"
        self.workbook.save(filename = "energy_per_round.xlsx")

    def add_energy_sent_to_battery(energy_sent):
        counter = counter + 1

module = Data_Collection_Module()

module.initialize_workbook()