from openpyxl import Workbook
import classes.event as event

class Data_Collection_Module:
    def __init__(self):
        self.counter = 1
        self.workbook = None
        self.ws = None
        self.filename = "energy_per_round.xlsx"
    
    def initialize_workbook(self):
        self.workbook = Workbook()
        self.ws = self.workbook.active
        self.ws.title = "Energy sent to battery"
        self.ws.cell(column=1,row=1).value = "Round"
        self.ws.cell(column=2,row=1).value = "Energy sent to battery by players in this round"
        self.ws.cell(column=3,row=1).value = "Event ID"
        self.ws.cell(column=4,row=1).value = "Event name"
        self.ws.cell(column=5,row=1).value = "Energy decision at end of this round"
        #player cells
        self.ws.cell(column=6,row=1).value = "Red"
        self.ws.cell(column=7,row=1).value = "Blue"
        self.ws.cell(column=8,row=1).value = "Green"
        self.ws.cell(column=9,row=1).value = "Yellow"
        self.workbook.save(filename = self.filename)

    def add_energy_sent_to_battery(self, energy_sent):
        self.counter = self.counter + 1
        self.ws.cell(column=1,row=self.counter).value = self.counter - 1
        self.ws.cell(column=2,row=self.counter).value = energy_sent
        self.workbook.save(filename = self.filename)

    def add_event(self, event):
        self.ws.cell(column=3,row=self.counter).value = event.id
        self.ws.cell(column=4,row=self.counter).value = event.name
        self.workbook.save(filename = self.filename)
    
    def add_energy_decision(self, energy_decision):
        # if the energy_decision object is None (which it is during excess outside of the final round in a season)
        # then we just save a string with text
        if energy_decision == None:
            self.ws.cell(column=5,row=self.counter).value = "We will save the excess energy for now."
        else:
            self.ws.cell(column=5,row=self.counter).value = energy_decision.flavor_text
        self.workbook.save(filename = self.filename)

    def add_player_energy(self, player_role, energy_to_battery):
        # loop through the columns for the players
        for column in range(6, 10):
            # if the value in the cell with player names set to lower is equal to the player's role save the data in that column at the counter row
            if self.ws.cell(column=column,row=1).value.lower() == player_role:
                # remember! here we set the row to counter+1, because we only update the counter after every player has inserted something
                self.ws.cell(column=column,row=self.counter+1).value = energy_to_battery


