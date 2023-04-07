import random
from openpyxl import load_workbook, Workbook

class Event:
    def __init__(self, id, name, flavor_text, thm_text, event_effect, season):
        self.id = id
        self.name = name
        self.flavor_text = flavor_text
        self.thm_text = thm_text
        self.event_effect = event_effect
        # season is a number - 0 for spring, 1 for summer, 2 for fall, 3 for winter
        self.season = season
        self.chosen_this_game = False

    def getEventText(self):
        return self.flavor_text

    # static method for the Event class to identify a random event
    @staticmethod
    def randomize_event(event_list, season):
        event_index = random.randint(0, len(event_list)-1)
        # this while loop breaks after there are no events with chosen_this_game == False
        # while loop to get events that are neither chosen nor in another season
        while event_list[event_index].chosen_this_game == True or event_list[event_index].season != season:
            event_index = random.randint(0, len(event_list)-1)
        # if event_list[event_index.season] == season:
            # get the event_list index item and set chosen this game to True
            # return the item
        # else:
            #Event.randomize_event()
        event_list[event_index].chosen_this_game = True        
        return event_list[event_index]
    
    # static method to get list of events that have passed in a season
    @staticmethod
    # take the game object's event_list and a season index as parameters
    def get_events(event_list, season):
        chosen_events = []
        for event in event_list:
            if event.chosen_this_game == True and event.season == season:
                chosen_events.append(event)
        # return the list of chosen events
        return chosen_events

# create the empty event list
event_list = []

# open workbook and worksheet
workbook = load_workbook("EVENTS.xlsx")
worksheet = workbook.active

# create event objects from Excel sheet and add them to the event_list
# go from 1 (which contains our headings) to stopping before 17, AKA at row 16
for i in range(1,17):
    new_event = Event( int(worksheet.cell(column=1,row=i+1).value), 
                      str(worksheet.cell(column=2,row=i+1).value), 
                      str(worksheet.cell(column=3,row=i+1).value),
                      str(worksheet.cell(column=4,row=i+1).value), 
                      int(worksheet.cell(column=5,row=i+1).value), 
                      int(worksheet.cell(column=6,row=i+1).value)
                      )
    event_list.append(new_event)

for i in event_list:
    print(i.name)
